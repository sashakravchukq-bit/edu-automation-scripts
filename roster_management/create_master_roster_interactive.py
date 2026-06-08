#!/usr/bin/env python3
"""
CREATE MASTER CLASS ROSTER - INTERACTIVE
=========================================

PURPOSE:
    Combines multiple class period rosters into a single master roster.
    Interactive version that guides teachers through the process step-by-step.

USAGE:
    python create_master_roster_interactive.py
    
    Or for help:
    python create_master_roster_interactive.py --help

AUTHOR: Educational Automation Library
VERSION: 2.0 (Interactive)
"""

import sys
import os
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# HELP DOCUMENTATION
# ============================================================================

HELP_TEXT = """
╔══════════════════════════════════════════════════════════════════════════╗
║              CREATE MASTER CLASS ROSTER - INTERACTIVE                     ║
╚══════════════════════════════════════════════════════════════════════════╝

PURPOSE:
    Combines multiple class period roster CSV files into one master Excel 
    file containing all students across all your class periods.

WHEN TO USE:
    • Beginning of semester to get organized
    • After roster changes to update your master list
    • When you need a complete student list for other scripts

BEFORE YOU START - YOU'LL NEED:
    
    CSV roster files exported from your gradebook system (one per period)
    
    Required information in each roster:
    • Student Name (column B, row 10+)
    • Grade/Year (column C, row 10+)
    • Student ID (column D, row 10+)
    • Period number (can be entered manually or auto-detected)

FILE FORMAT:
    Your roster CSVs should look like standard gradebook exports:
    
    Row 1-2:   Teacher info, course info, dates
    Row 7:     Column headers (Indicators, Student Name, Grade, etc.)
    Row 10+:   Student data starts here
    
    Example:
    "Teacher:","Smith, Jane",...
    "Course:","AP Computer Science Principles","Period: 1"
    ...
    "Indicators","Student Name","Grade","Student Number",...
    
    "","Garcia, Maria","10","12345",...
    "","Johnson, Robert","11","12346",...

OUTPUT:
    Excel file with columns:
    • Student Name
    • Grade
    • Student ID
    • Period
    • Course
    
    Features:
    • Sorted by period, then alphabetically
    • Color-coded header row
    • Preserves all student information
    • Ready to use with other automation scripts

WORKFLOW:
    1. Export roster CSVs from your gradebook (one per period)
    2. Save them all in one folder
    3. Run this script
    4. Enter file paths when prompted
    5. Get one master roster Excel file

TIPS:
    • Keep roster files in a consistent folder structure
    • Use clear file naming: COURSE_Semester_Period_#_Roster.csv
    • You can drag-and-drop files into terminal for instant paths
    • Press Enter to accept default values shown in [brackets]

═══════════════════════════════════════════════════════════════════════════

Press Enter to continue or Ctrl+C to exit...
"""

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def print_header():
    """Print a nice header for the script."""
    print("\n" + "="*78)
    print("  CREATE MASTER CLASS ROSTER - Interactive Mode")
    print("="*78 + "\n")

def print_section(title):
    """Print a section divider."""
    print(f"\n{'─'*78}")
    print(f"  {title}")
    print('─'*78)

def get_input(prompt, default=None, required=True):
    """Get user input with optional default value."""
    if default is not None:
        full_prompt = f"{prompt} [{default}]: "
    else:
        full_prompt = f"{prompt}: "
    
    while True:
        value = input(full_prompt).strip()
        
        if not value and default is not None:
            return str(default)
        
        if required and not value:
            print("  ⚠ This field is required. Please enter a value.")
            continue
        
        return value

def get_file_path(prompt, default=None, must_exist=True):
    """Get a file path from user with validation."""
    while True:
        path = get_input(prompt, default)
        
        path = os.path.expanduser(path)
        path = os.path.abspath(path)
        path = path.strip('"').strip("'")
        
        if must_exist and not os.path.exists(path):
            print(f"  ⚠ File not found: {path}")
            print(f"  Please check the path and try again.")
            continue
        
        return path

def get_yes_no(prompt, default='y'):
    """Get yes/no response from user."""
    default_text = 'Y/n' if default.lower() == 'y' else 'y/N'
    response = input(f"{prompt} [{default_text}]: ").strip().lower()
    
    if not response:
        response = default
    
    return response in ['y', 'yes']

def extract_student_data(file_path, period, course):
    """
    Extract student information from a roster CSV file.
    
    Args:
        file_path: Path to the CSV roster file
        period: Class period number
        course: Course code/name
        
    Returns:
        list: List of dictionaries containing student information
    """
    students = []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        reader = csv.reader(lines)
        roster_data = list(reader)
        
        # Student data typically starts at row 9 (index 9)
        for idx in range(9, len(roster_data)):
            row = roster_data[idx]
            
            if len(row) > 3:
                student_name = row[1] if len(row) > 1 else ""
                grade = row[2] if len(row) > 2 else ""
                student_id = row[3] if len(row) > 3 else ""
                
                # Only add if we have valid data
                if student_name and student_name.strip() and student_id and student_id.strip():
                    students.append({
                        'Student Name': student_name.strip(),
                        'Grade': grade.strip(),
                        'Student ID': student_id.strip(),
                        'Period': period,
                        'Course': course
                    })
    
    except Exception as e:
        print(f"  ⚠ Error reading file: {e}")
        return []
    
    return students

# ============================================================================
# MAIN INTERACTIVE SCRIPT
# ============================================================================

def main():
    """Main interactive workflow."""
    
    # Check for help flag
    if '--help' in sys.argv or '-h' in sys.argv:
        print(HELP_TEXT)
        input()
        return
    
    print_header()
    
    print("Welcome! This script will help you create a master class roster.")
    print("You can press Ctrl+C at any time to cancel.\n")
    
    if get_yes_no("Would you like to see detailed help before starting?", default='n'):
        print(HELP_TEXT)
        input("Press Enter to continue...")
    
    # ========================================================================
    # STEP 1: Collect Roster Files
    # ========================================================================
    
    print_section("Step 1: Collect Roster Files")
    print("I'll ask you for each class period roster CSV file.")
    print("These should be exported from your gradebook system.\n")
    
    rosters = []
    
    # Ask how many periods
    while True:
        try:
            num_periods = int(get_input("How many class periods do you have?", default="4"))
            if num_periods > 0:
                break
            print("  ⚠ Please enter a number greater than 0.")
        except ValueError:
            print("  ⚠ Please enter a valid number.")
    
    print(f"\n  ✓ Will collect roster information for {num_periods} period(s)\n")
    
    # Collect each roster
    for i in range(num_periods):
        print(f"\n── Period {i+1} ──")
        
        has_file = get_yes_no(f"Do you have a roster file for this period?", default='y')
        
        if not has_file:
            print(f"  Skipping period {i+1}")
            continue
        
        # Get period number
        period_num = int(get_input("  What is the period number?", default=str(i+1)))
        
        # Get course code
        print("  What course is this? (e.g., APCSP, ACS1, Algebra)")
        course_code = get_input("  Course code", default="APCSP")
        
        # Get file path
        default_path = f"./Period_{period_num}_Roster.csv"
        file_path = get_file_path(
            "  Enter path to roster CSV file",
            default=default_path if os.path.exists(default_path) else None
        )
        
        # Test read the file
        test_students = extract_student_data(file_path, period_num, course_code)
        
        if test_students:
            print(f"  ✓ Successfully read {len(test_students)} student(s)")
            rosters.append((file_path, period_num, course_code))
        else:
            print(f"  ⚠ Warning: No students found in this file")
            if get_yes_no("  Add it anyway?", default='n'):
                rosters.append((file_path, period_num, course_code))
    
    if not rosters:
        print("\n  ⚠ No roster files added. Nothing to process!")
        return
    
    print(f"\n  ✓ Collected {len(rosters)} roster file(s)")
    
    # ========================================================================
    # STEP 2: Output Location
    # ========================================================================
    
    print_section("Step 2: Output Location")
    print("Where should I save the master roster?\n")
    
    default_output = "./Master_Class_Roster.xlsx"
    output_path = get_file_path(
        "Enter output file path",
        default=default_output,
        must_exist=False
    )
    
    # Ensure .xlsx extension
    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'
    
    print(f"\n  ✓ Will save master roster to: {output_path}")
    
    # ========================================================================
    # STEP 3: Confirm and Process
    # ========================================================================
    
    print_section("Step 3: Ready to Process")
    print("\nSummary:")
    print(f"  • Number of period rosters: {len(rosters)}")
    print(f"  • Output file: {output_path}")
    print()
    
    if not get_yes_no("Ready to create master roster?", default='y'):
        print("\n  Cancelled. No files were created.")
        return
    
    # ========================================================================
    # PROCESSING
    # ========================================================================
    
    print_section("Processing Rosters")
    
    all_students = []
    
    for file_path, period, course in rosters:
        print(f"\nProcessing {course} Period {period}...")
        
        students = extract_student_data(file_path, period, course)
        
        if students:
            print(f"  ✓ Added {len(students)} student(s)")
            all_students.extend(students)
        else:
            print(f"  ⚠ No students found in this file")
    
    if not all_students:
        print("\n  ⚠ No students extracted from any files!")
        print("  Please check your roster file format and try again.")
        return
    
    print(f"\n  ✓ Total students collected: {len(all_students)}")
    
    # Sort by period, then by name
    all_students.sort(key=lambda x: (x['Period'], x['Student Name']))
    
    # ========================================================================
    # CREATE EXCEL WORKBOOK
    # ========================================================================
    
    print("\nCreating Excel workbook...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Roster"
    
    # Headers
    headers = ['Student Name', 'Grade', 'Student ID', 'Period', 'Course']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add student data
    for row_idx, student in enumerate(all_students, start=2):
        ws.cell(row=row_idx, column=1, value=student['Student Name'])
        ws.cell(row=row_idx, column=2, value=student['Grade'])
        ws.cell(row=row_idx, column=3, value=student['Student ID'])
        ws.cell(row=row_idx, column=4, value=student['Period'])
        ws.cell(row=row_idx, column=5, value=student['Course'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    
    # Save
    wb.save(output_path)
    
    # ========================================================================
    # SUMMARY
    # ========================================================================
    
    print_section("Master Roster Created!")
    
    print(f"\n📊 Summary:")
    print(f"  • Total students: {len(all_students)}")
    
    # Count by period
    period_counts = {}
    for student in all_students:
        key = f"Period {student['Period']} ({student['Course']})"
        period_counts[key] = period_counts.get(key, 0) + 1
    
    print(f"\n📊 By Period:")
    for period, count in sorted(period_counts.items()):
        print(f"  • {period}: {count} students")
    
    print(f"\n💾 Master roster saved to:")
    print(f"  {output_path}")
    
    print("\n" + "="*78)
    print("Next steps:")
    print("  1. Open the Excel file to verify all students are present")
    print("  2. Keep this file - other scripts will use it as input")
    print("  3. Update it whenever roster changes occur")
    print("="*78 + "\n")

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Cancelled by user. No files were created.")
        sys.exit(0)
    except Exception as e:
        print(f"\n  ⚠ ERROR: {e}")
        print(f"\n  If you need help, run: python {sys.argv[0]} --help")
        sys.exit(1)
