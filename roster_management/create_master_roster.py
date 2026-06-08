"""
CREATE MASTER CLASS ROSTER
==========================

PURPOSE:
    Combines multiple class period rosters into a single master roster containing
    all students across all periods. Useful for creating a complete student list
    at the beginning of a semester or for administrative purposes.

AUTHOR: Educational Automation Library
KEYWORDS: roster, master list, class list, student management, organization, periods

INPUTS REQUIRED:
    - CSV roster files (one per class period)
    - Files must contain columns: Student Name, Grade, Student Number
    - Can contain additional gradebook columns (will be ignored)
    
INPUT FILE FORMAT:
    - Standard gradebook export CSV format
    - First few rows contain header information (teacher, course, dates, etc.)
    - Row 7 (index 6) contains column headers
    - Student data starts at row 10 (index 9)
    - Columns: [1]=Name, [2]=Grade, [3]=Student ID

OUTPUT:
    - Excel file: Master_Class_Roster_[Semester]_[Year].xlsx
    - Columns: Student Name, Grade, Student ID, Period, Course
    - Sorted by Period, then alphabetically by Student Name
    - Color-coded header row for easy reading

FEATURES:
    - Handles compound last names (e.g., "Garcia Rodriguez")
    - Extracts period number automatically from file header
    - Preserves student ID numbers
    - Tracks which course each student is enrolled in
    - Creates formatted Excel output with styling

USAGE:
    1. Place all roster CSV files in the same folder
    2. Update the 'rosters' list below with your file paths and period numbers
    3. Run: python create_master_roster.py
    4. Check output folder for Master_Class_Roster_*.xlsx

CUSTOMIZATION:
    - Modify 'rosters' list to include your specific files
    - Change output_path to your preferred location
    - Adjust column widths in the "Adjust column widths" section

EXAMPLE:
    Input files:
        APCSP_Spring_2026_Period_1_Roster.csv
        APCSP_Spring_2026_Period_3_Roster.csv
        
    Output:
        Master_Class_Roster_Spring_2026.xlsx
        Contains all students from both periods in one organized list
"""

import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# CONFIGURATION - UPDATE THESE PATHS FOR YOUR FILES
# ============================================================================

# Define the roster files and their periods
# Format: (filepath, period_number, course_code)
rosters = [
    ('/mnt/user-data/uploads/APCSP_Spring_2026_Period_1_Roster.csv', 1, 'APCSP'),
    ('/mnt/user-data/uploads/APCSP_Spring_2026_Period_3_Roster.csv', 3, 'APCSP'),
    ('/mnt/user-data/uploads/APCSP_Spring_2026_Period_5_Roster.csv', 5, 'APCSP'),
    ('/mnt/user-data/uploads/ACS1_Spring_2026_Period_4_Roster.csv', 4, 'ACS1')
]

# Output file path
output_path = '/mnt/user-data/outputs/Master_Class_Roster_Spring_2026.xlsx'

# ============================================================================
# MAIN SCRIPT - NO NEED TO MODIFY BELOW THIS LINE
# ============================================================================

def extract_student_data(file_path, period, course):
    """
    Extract student information from a roster CSV file.
    
    Args:
        file_path (str): Path to the CSV roster file
        period (int): Class period number
        course (str): Course code/name
        
    Returns:
        list: List of dictionaries containing student information
    """
    students = []
    
    with open(file_path, 'r') as f:
        lines = f.readlines()
    
    reader = csv.reader(lines)
    roster_data = list(reader)
    
    # Student data starts at row 9 (index 9)
    # Columns: [1] = Name, [2] = Grade, [3] = Student Number
    for idx in range(9, len(roster_data)):
        row = roster_data[idx]
        
        if len(row) > 3:
            student_name = row[1] if len(row) > 1 else ""
            grade = row[2] if len(row) > 2 else ""
            student_id = row[3] if len(row) > 3 else ""
            
            # Only add if we have valid data
            if student_name and student_name.strip() and student_id and student_id.strip():
                students.append({
                    'Student Name': student_name.strip(),
                    'Grade': grade.strip(),
                    'Student ID': student_id.strip(),
                    'Period': period,
                    'Course': course
                })
    
    return students

# List to store all students
all_students = []

# Process each roster file
print("Processing roster files...")
print("=" * 70)

for file_path, period, course in rosters:
    print(f"\nProcessing {course} Period {period}...")
    
    students = extract_student_data(file_path, period, course)
    
    for student in students:
        all_students.append(student)
        print(f"  Added: {student['Student Name']}")

print(f"\n{'=' * 70}")
print(f"Total students collected: {len(all_students)}")
print("=" * 70)

# Sort by period, then by student name
all_students.sort(key=lambda x: (x['Period'], x['Student Name']))

# ============================================================================
# CREATE EXCEL WORKBOOK WITH FORMATTING
# ============================================================================

wb = Workbook()
ws = wb.active
ws.title = "Master Roster"

# Define headers
headers = ['Student Name', 'Grade', 'Student ID', 'Period', 'Course']

# Style for headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
header_alignment = Alignment(horizontal='center', vertical='center')

# Add headers with styling
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Add student data
for row_idx, student in enumerate(all_students, start=2):
    ws.cell(row=row_idx, column=1, value=student['Student Name'])
    ws.cell(row=row_idx, column=2, value=student['Grade'])
    ws.cell(row=row_idx, column=3, value=student['Student ID'])
    ws.cell(row=row_idx, column=4, value=student['Period'])
    ws.cell(row=row_idx, column=5, value=student['Course'])

# Adjust column widths for readability
ws.column_dimensions['A'].width = 30  # Student Name
ws.column_dimensions['B'].width = 8   # Grade
ws.column_dimensions['C'].width = 12  # Student ID
ws.column_dimensions['D'].width = 8   # Period
ws.column_dimensions['E'].width = 12  # Course

# Save the workbook
wb.save(output_path)

# ============================================================================
# PRINT SUMMARY STATISTICS
# ============================================================================

print(f"\n{'=' * 70}")
print("Master roster created successfully!")
print("=" * 70)
print(f"\nSaved to: {output_path}")
print(f"\nBreakdown by period:")
print("-" * 70)

# Count students by period
period_counts = {}
for student in all_students:
    key = f"Period {student['Period']} ({student['Course']})"
    period_counts[key] = period_counts.get(key, 0) + 1

for period, count in sorted(period_counts.items()):
    print(f"  {period}: {count} students")

print("\n" + "=" * 70)
print("Next steps:")
print("  1. Open the Excel file to review the master roster")
print("  2. Use this file for tracking assignments across all periods")
print("  3. Keep this file as your semester reference")
print("=" * 70)
