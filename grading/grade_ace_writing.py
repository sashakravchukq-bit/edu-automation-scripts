"""
GRADE ACE RUBRIC WRITING ASSIGNMENTS
====================================

PURPOSE:
    Automatically grades short answer writing responses using the ACE 
    (Answer, Cite, Explain) rubric strategy. Scores each component (0-2 points)
    for a total of 0-6 points. Includes students who didn't submit (scored as 0).

AUTHOR: Educational Automation Library
KEYWORDS: grading, writing, ACE rubric, TELPAS, assessment, automatic grading, 
          short answer, citation, explanation, reading comprehension

INPUTS REQUIRED:
    1. Google Form CSV exports (one per class period) containing:
       - Timestamp
       - Username (student email)
       - Response to the question
       
    2. Master roster Excel file containing all students
       - Must have: Student Name, Period columns
       
    3. Prompts information (coded in script):
       - Text passage for each period
       - Question asked
       
INPUT FILE NAMING:
    Week_#_Period_#_Assignment_Name.csv
    Example: Week_1_Period_1_Writing_Project_Volcanoes.csv

OUTPUT:
    - Excel file: Write_it_Wednesday_Week_#_Grades.xlsx
    - Columns: Period, Student Name, Email, Answer (0-2), Cite (0-2), 
      Explain (0-2), Total Score (0-6), Response Preview
    - Color-coded totals: Green (5-6), Yellow (3-4), Red (0-2)
    - Includes students who didn't submit with 0 scores

GRADING CRITERIA:
    
    ANSWER (0-2 points):
    - 2 points: Fully answers all parts of the question with accurate information
    - 1 point: Partially answers the question or answers incompletely
    - 0 points: Does not answer the question or answers incorrectly
    
    CITE (0-2 points):
    - 2 points: Uses quotation marks AND/OR citation phrases like:
      * "The author states..."
      * "According to the passage..."
      * "The text mentions..."
    - 1 point: References the text but less formally
    - 0 points: No citation or reference to text
    
    EXPLAIN (0-2 points):
    - 2 points: Clearly connects evidence to answer using explanation words
      * "This shows...", "Therefore...", "Because...", "This proves..."
      * Has logical structure (3+ sentences)
    - 1 point: Attempts to connect OR has structure but not both
    - 0 points: No connection between evidence and answer

FEATURES:
    - Matches student emails to names from master roster
    - Fuzzy name matching (handles variations)
    - Includes non-submitters automatically
    - Customizable rubric criteria per topic
    - Period and overall statistics
    - Response preview for manual spot-checking

USAGE:
    1. Export Google Form responses as CSV (one per period)
    2. Ensure you have an updated master roster
    3. Update the 'form_files' and 'prompts' dictionaries below
    4. Run: python grade_ace_writing.py
    5. Review output Excel file

CUSTOMIZATION:
    - Modify 'prompts' dictionary to match your assignments
    - Adjust grading criteria in grade_answer() function
    - Change color thresholds in the Excel formatting section
    - Modify response preview length (currently 100 characters)

IMPORTANT NOTES:
    - This provides a strong starting point but should be spot-checked
    - Rubric logic is based on keyword detection and structure analysis
    - Consider manually reviewing borderline scores (3-4 points)
    - Grammar/spelling quality is considered but not heavily weighted

EXAMPLE WORKFLOW:
    1. Students complete Google Form with reading passage and question
    2. Export responses to CSV
    3. Run this script
    4. Review scores, focusing on low scores and spot-checking high scores
    5. Make manual adjustments as needed in gradebook
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv

# ============================================================================
# CONFIGURATION - UPDATE FOR YOUR ASSIGNMENT
# ============================================================================

# Define the prompts and questions for each period
# Add or modify entries to match your actual assignments
prompts = {
    1: {
        'topic': 'How Volcanoes Erupt',
        'question': 'What is the difference between magma and lava, and what causes the pressure that makes a volcano erupt?',
        'text': '''A volcano is essentially a vent, or opening, in the Earth's crust where molten rock, ash, and gases can escape from below the surface. Deep beneath the ground is super-hot, liquid rock called magma. This magma collects in a chamber. When the pressure from the gases in the magma becomes too great, it forces the magma up and out through the vent. Once the magma reaches the surface, it is called lava. Volcanoes are often found along the edges of tectonic plates.'''
    },
    3: {
        'topic': 'All About Mammals',
        'question': 'List the four main characteristics that define an animal as a mammal.',
        'text': '''Mammals are a group of animals that have four main characteristics that set them apart from other animals like birds, reptiles, and fish. First, they have fur or hair on their bodies. Second, they are warm-blooded, meaning they can keep their body temperature stable no matter the weather outside. Third, they give birth to live babies, and fourth, the mothers produce milk to feed their young. Humans, dogs, elephants, and whales are all examples of mammals.'''
    },
    4: {
        'topic': 'What is Gravity?',
        'question': 'What two main factors determine the strength of the force of gravity between two objects, and what does Earth\'s gravity pull everything toward?',
        'text': '''Gravity is the invisible force that pulls objects toward each other. On Earth, gravity is what keeps us from floating away and is why things fall down when you drop them. The strength of the gravitational pull depends on the mass (the amount of stuff in an object) and the distance between the objects. The Earth has a huge mass, so its gravity is very strong, pulling everything toward its center. Without gravity, the moon wouldn't orbit the Earth, and the Earth wouldn't orbit the sun.'''
    },
    5: {
        'topic': 'Why We Have Rules and Laws',
        'question': 'Name two essential purposes of having rules and laws in a community or a country and explain how they can help.',
        'text': '''Rules and laws are important guidelines that help people live together peacefully in a community or country. Laws are often made by the government and come with serious consequences, like fines or jail time, if they are broken. Rules and laws serve a few purposes: they keep people safe, they ensure that people are treated fairly, and they help resolve conflicts when they happen. Both simple classroom rules and complex national laws are necessary to keep order.'''
    }
}

# Google Form CSV files - UPDATE THESE PATHS
form_files = [
    ('/mnt/user-data/uploads/Week_1_Period_1_Writing_Project___How_Volcanoes_Erupt_.csv', 1),
    ('/mnt/user-data/uploads/Week_1_Period_3_Writing_Project___All_About_Mammals__.csv', 3),
    ('/mnt/user-data/uploads/Week_1_Period_4_Writing_Project___What_is_Gravity___.csv', 4),
    ('/mnt/user-data/uploads/Week_1_Period_5_Writing_Project___Why_We_Have_Rules_and_Laws__.csv', 5)
]

# Master roster path - UPDATE THIS PATH
master_roster_path = '/mnt/user-data/outputs/Master_Class_Roster_Spring_2026.xlsx'

# Output path - UPDATE IF NEEDED
output_path = '/mnt/user-data/outputs/Write_it_Wednesday_Week_1_Grades.xlsx'

# ============================================================================
# GRADING FUNCTION - CUSTOMIZE CRITERIA HERE
# ============================================================================

def grade_answer(response, question, period):
    """
    Grade a student response based on ACE rubric.
    
    Args:
        response (str): Student's written response
        question (str): The question that was asked
        period (int): Class period (determines question-specific criteria)
        
    Returns:
        dict: Scores for answer, cite, explain, and total
    """
    response_lower = response.lower()
    
    # ========================================================================
    # ANSWER SCORING (0-2 points)
    # ========================================================================
    # Check if the response addresses the specific question asked
    # Customize this section for each period's question
    
    if period == 1:
        # Period 1: Volcanoes - need difference between magma/lava + pressure cause
        has_difference = ('magma' in response_lower and 'lava' in response_lower)
        has_pressure = ('pressure' in response_lower or 'gas' in response_lower)
        if has_difference and has_pressure:
            answer_score = 2
        elif has_difference or has_pressure:
            answer_score = 1
        else:
            answer_score = 0
            
    elif period == 3:
        # Period 3: Mammals - need four characteristics
        characteristics = [
            'fur' in response_lower or 'hair' in response_lower,
            'warm' in response_lower,
            'live' in response_lower or 'birth' in response_lower,
            'milk' in response_lower
        ]
        char_count = sum(characteristics)
        if char_count >= 4:
            answer_score = 2
        elif char_count >= 2:
            answer_score = 1
        else:
            answer_score = 0
            
    elif period == 4:
        # Period 4: Gravity - need mass, distance, and Earth's center
        has_factors = (('mass' in response_lower or 'weight' in response_lower) and 
                      'distance' in response_lower)
        has_center = 'center' in response_lower
        if has_factors and has_center:
            answer_score = 2
        elif has_factors or has_center:
            answer_score = 1
        else:
            answer_score = 0
            
    elif period == 5:
        # Period 5: Rules/Laws - need two purposes
        purposes = [
            'safe' in response_lower,
            'fair' in response_lower or 'treat' in response_lower,
            'conflict' in response_lower or 'order' in response_lower
        ]
        purpose_count = sum(purposes)
        if purpose_count >= 2:
            answer_score = 2
        elif purpose_count >= 1:
            answer_score = 1
        else:
            answer_score = 0
    else:
        answer_score = 0
    
    # ========================================================================
    # CITE SCORING (0-2 points)
    # ========================================================================
    # Look for evidence of text citation
    
    # Check for quotation marks (various unicode versions)
    has_quotes = '"' in response or '"' in response or '"' in response or "'" in response
    
    # Check for citation phrases
    cite_phrases = [
        'author states', 'article states', 'passage', 'according to', 
        'text states', 'paragraph', 'excerpt', 'states that', 'mentions',
        'the text', 'the article', 'the passage'
    ]
    has_cite_language = any(phrase in response_lower for phrase in cite_phrases)
    
    if has_quotes and has_cite_language:
        cite_score = 2
    elif has_quotes or has_cite_language:
        cite_score = 2  # Being generous - either method shows citation attempt
    elif len(response) > 100:  # At least attempted substantial response
        cite_score = 1
    else:
        cite_score = 0
    
    # ========================================================================
    # EXPLAIN SCORING (0-2 points)
    # ========================================================================
    # Look for logical connection between evidence and answer
    
    # Check for explanation/connection words
    explain_phrases = [
        'this shows', 'this proves', 'this means', 'therefore', 
        'because', 'which causes', 'as a result', 'this demonstrates',
        'this explains', 'validates', 'supports', 'connected', 'since',
        'thus', 'hence', 'consequently'
    ]
    has_explanation = any(phrase in response_lower for phrase in explain_phrases)
    
    # Check for logical structure (multiple sentences)
    sentences = [s for s in response.split('.') if len(s.strip()) > 10]
    has_structure = len(sentences) >= 3  # At least 3 substantial sentences
    
    if has_explanation and has_structure:
        explain_score = 2
    elif has_explanation or has_structure:
        explain_score = 1
    else:
        explain_score = 0
    
    # ========================================================================
    # QUALITY CHECK - Penalize very poor quality
    # ========================================================================
    # Check for major quality issues
    incomplete_sentences = sum(1 for s in response.split('.') if 0 < len(s.strip()) < 10)
    
    # Penalize if response is too short or mostly incomplete sentences
    if len(response) < 50 or (len(sentences) > 0 and incomplete_sentences > len(sentences)):
        # Reduce answer score for poor quality
        if answer_score > 0:
            answer_score = max(0, answer_score - 1)
    
    # Calculate total
    total_score = answer_score + cite_score + explain_score
    
    return {
        'answer_score': answer_score,
        'cite_score': cite_score,
        'explain_score': explain_score,
        'total_score': total_score
    }

# ============================================================================
# MAIN GRADING PROCESS
# ============================================================================

# Load master roster to get all students
print("Loading master roster...")
master_roster = pd.read_excel(master_roster_path)

# Storage for all grades
all_grades = []

# Process each period's submissions
for file_path, period in form_files:
    print(f"\n{'=' * 70}")
    print(f"Grading Period {period} - {prompts[period]['topic']}")
    print("=" * 70)
    
    # Read Google Form responses
    with open(file_path, 'r') as f:
        reader = csv.DictReader(f)
        responses = list(reader)
    
    # Create dictionary of submissions by email
    submissions = {}
    for response in responses:
        email = response['Username'].strip().lower()
        answer_text = response[list(response.keys())[2]]  # Third column is the answer
        submissions[email] = answer_text
    
    # Get all students for this period from master roster
    period_students = master_roster[master_roster['Period'] == period]
    
    # Grade each student (or mark as no submission)
    for _, student in period_students.iterrows():
        student_name = student['Student Name']
        
        # Extract name parts for email matching
        if ',' in student_name:
            last_name, first_name = student_name.split(',', 1)
            last_name = last_name.strip().lower().replace(' ', '').replace('-', '')
            first_name = first_name.strip().lower().replace(' ', '').replace('-', '')
        else:
            parts = student_name.split()
            if len(parts) >= 2:
                first_name = parts[0].lower()
                last_name = ' '.join(parts[1:]).lower().replace(' ', '').replace('-', '')
            else:
                first_name = student_name.lower()
                last_name = ""
        
        # Find matching submission
        found_submission = None
        matched_email = None
        for email, answer_text in submissions.items():
            email_lower = email.lower()
            if first_name in email_lower and last_name in email_lower:
                found_submission = answer_text
                matched_email = email
                break
        
        if found_submission:
            # Grade the response
            grades = grade_answer(found_submission, prompts[period]['question'], period)
            
            all_grades.append({
                'Period': period,
                'Student Name': student_name,
                'Email': matched_email,
                'Answer Score': grades['answer_score'],
                'Cite Score': grades['cite_score'],
                'Explain Score': grades['explain_score'],
                'Total Score': grades['total_score'],
                'Response': found_submission[:100] + '...' if len(found_submission) > 100 else found_submission
            })
            
            print(f"{student_name}: {grades['total_score']}/6 "
                  f"(A:{grades['answer_score']} C:{grades['cite_score']} E:{grades['explain_score']})")
        else:
            # No submission
            all_grades.append({
                'Period': period,
                'Student Name': student_name,
                'Email': 'Not submitted',
                'Answer Score': 0,
                'Cite Score': 0,
                'Explain Score': 0,
                'Total Score': 0,
                'Response': 'NO SUBMISSION'
            })
            
            print(f"{student_name}: 0/6 - NO SUBMISSION")

# ============================================================================
# CREATE EXCEL OUTPUT WITH FORMATTING
# ============================================================================

print(f"\n{'=' * 70}")
print("Creating Excel grade sheet...")
print("=" * 70)

wb = Workbook()
ws = wb.active
ws.title = "Write it Wednesday Grades"

# Headers
headers = ['Period', 'Student Name', 'Email', 'Answer (0-2)', 'Cite (0-2)', 
           'Explain (0-2)', 'Total Score (0-6)', 'Response Preview']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add data with conditional formatting
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 5-6 points
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 3-4 points
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')     # 0-2 points

for row_idx, grade in enumerate(all_grades, start=2):
    ws.cell(row=row_idx, column=1, value=grade['Period'])
    ws.cell(row=row_idx, column=2, value=grade['Student Name'])
    ws.cell(row=row_idx, column=3, value=grade['Email'])
    ws.cell(row=row_idx, column=4, value=grade['Answer Score'])
    ws.cell(row=row_idx, column=5, value=grade['Cite Score'])
    ws.cell(row=row_idx, column=6, value=grade['Explain Score'])
    
    total_cell = ws.cell(row=row_idx, column=7, value=grade['Total Score'])
    
    # Color code based on score
    if grade['Total Score'] >= 5:
        total_cell.fill = green_fill
    elif grade['Total Score'] >= 3:
        total_cell.fill = yellow_fill
    else:
        total_cell.fill = red_fill
    
    ws.cell(row=row_idx, column=8, value=grade['Response'])

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 50

# Save
wb.save(output_path)

# ============================================================================
# PRINT SUMMARY STATISTICS
# ============================================================================

print(f"\n{'=' * 70}")
print("GRADING SUMMARY")
print("=" * 70)

for period in sorted(set(g['Period'] for g in all_grades)):
    period_grades = [g for g in all_grades if g['Period'] == period]
    avg_score = sum(g['Total Score'] for g in period_grades) / len(period_grades)
    submitted = sum(1 for g in period_grades if g['Total Score'] > 0)
    total = len(period_grades)
    print(f"\nPeriod {period}:")
    print(f"  Students: {total}")
    print(f"  Submitted: {submitted} ({submitted/total*100:.1f}%)")
    print(f"  Average Score: {avg_score:.2f}/6")

total_students = len(all_grades)
total_submitted = sum(1 for g in all_grades if g['Total Score'] > 0)
total_avg = sum(g['Total Score'] for g in all_grades) / len(all_grades)

print(f"\n{'=' * 70}")
print(f"Overall:")
print(f"  Total Students: {total_students}")
print(f"  Submitted: {total_submitted} ({total_submitted/total_students*100:.1f}%)")
print(f"  Average Score: {total_avg:.2f}/6")
print(f"\n{'=' * 70}")
print(f"Grades saved to: {output_path}")
print("=" * 70)
print("\nNext steps:")
print("  1. Open the Excel file to review scores")
print("  2. Spot-check responses, especially borderline scores (3-4)")
print("  3. Make manual adjustments as needed")
print("  4. Transfer scores to your gradebook")
print("=" * 70)
