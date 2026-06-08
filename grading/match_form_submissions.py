import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load the master roster
master_roster = pd.read_excel('/mnt/user-data/outputs/Master_Class_Roster_Spring_2026.xlsx')

# Define the Google Form files and their periods
form_files = [
    ('/mnt/user-data/uploads/Week_1_Period_1_Writing_Project___How_Volcanoes_Erupt_.csv', 1),
    ('/mnt/user-data/uploads/Week_1_Period_3_Writing_Project___All_About_Mammals__.csv', 3),
    ('/mnt/user-data/uploads/Week_1_Period_4_Writing_Project___What_is_Gravity___.csv', 4),
    ('/mnt/user-data/uploads/Week_1_Period_5_Writing_Project___Why_We_Have_Rules_and_Laws__.csv', 5)
]

# Dictionary to store who completed by period
completed_by_period = {}

# Process each Google Form file
for file_path, period in form_files:
    print(f"\nProcessing Period {period} submissions...")
    
    # Read the form responses
    form_df = pd.read_csv(file_path)
    
    # Extract usernames (email addresses)
    if 'Username' in form_df.columns:
        submitted_emails = set(form_df['Username'].str.lower().str.strip())
        completed_by_period[period] = submitted_emails
        print(f"  Found {len(submitted_emails)} submissions")
    else:
        print(f"  ERROR: No 'Username' column found in {file_path}")
        completed_by_period[period] = set()

# Create a completion report
print("\n" + "="*60)
print("Creating completion report...")
print("="*60)

# Prepare data for the report
report_data = []

for _, student in master_roster.iterrows():
    period = student['Period']
    student_name = student['Student Name']
    grade = student['Grade']
    student_id = student['Student ID']
    course = student['Course']
    
    # Try to match student to email
    # Extract first and last name
    if ',' in student_name:
        last_name, first_name = student_name.split(',', 1)
        last_name = last_name.strip()
        first_name = first_name.strip()
    else:
        # Handle names without comma
        parts = student_name.split()
        if len(parts) >= 2:
            first_name = parts[0]
            last_name = ' '.join(parts[1:])
        else:
            first_name = student_name
            last_name = ""
    
    # Check if this student's period had submissions
    if period in completed_by_period:
        submitted_emails = completed_by_period[period]
        
        # Try to find a matching email
        # Common pattern: firstname.lastname###@uisd.net
        found_match = False
        
        for email in submitted_emails:
            # Check if email contains the student's name parts
            email_lower = email.lower()
            first_lower = first_name.lower().replace(' ', '').replace('-', '')
            last_lower = last_name.lower().replace(' ', '').replace('-', '')
            
            # Try various matching patterns
            if first_lower in email_lower and last_lower in email_lower:
                found_match = True
                break
        
        completed = "Yes" if found_match else "No"
    else:
        completed = "No Form"
    
    report_data.append({
        'Student Name': student_name,
        'Grade': grade,
        'Student ID': student_id,
        'Period': period,
        'Course': course,
        'Write it Wednesday': completed
    })

# Create Excel workbook for report
wb = Workbook()
ws = wb.active
ws.title = "Write it Wednesday Report"

# Add headers with styling
headers = ['Student Name', 'Grade', 'Student ID', 'Period', 'Course', 'Write it Wednesday']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add data with conditional formatting
yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for row_idx, student in enumerate(report_data, start=2):
    ws.cell(row=row_idx, column=1, value=student['Student Name'])
    ws.cell(row=row_idx, column=2, value=student['Grade'])
    ws.cell(row=row_idx, column=3, value=student['Student ID'])
    ws.cell(row=row_idx, column=4, value=student['Period'])
    ws.cell(row=row_idx, column=5, value=student['Course'])
    
    completion_cell = ws.cell(row=row_idx, column=6, value=student['Write it Wednesday'])
    
    # Color code the completion status
    if student['Write it Wednesday'] == 'Yes':
        completion_cell.fill = yes_fill
    elif student['Write it Wednesday'] == 'No':
        completion_cell.fill = no_fill

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 20

# Save the workbook
output_path = '/mnt/user-data/outputs/Write_it_Wednesday_Week_1_Report.xlsx'
wb.save(output_path)

# Print summary statistics
print("\nCompletion Summary by Period:")
for period in sorted(completed_by_period.keys()):
    period_students = [s for s in report_data if s['Period'] == period]
    completed = len([s for s in period_students if s['Write it Wednesday'] == 'Yes'])
    total = len(period_students)
    percentage = (completed / total * 100) if total > 0 else 0
    print(f"  Period {period}: {completed}/{total} ({percentage:.1f}%)")

total_students = len(report_data)
total_completed = len([s for s in report_data if s['Write it Wednesday'] == 'Yes'])
total_percentage = (total_completed / total_students * 100) if total_students > 0 else 0
print(f"\nOverall: {total_completed}/{total_students} ({total_percentage:.1f}%)")

print(f"\nReport saved to: {output_path}")
