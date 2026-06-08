#!/usr/bin/env python3
"""
INTERACTIVE ACE RUBRIC WRITING GRADER
=====================================

PURPOSE:
    User-friendly version of the ACE rubric grader with interactive prompts
    that guide teachers through the grading process step-by-step.

USAGE:
    python grade_ace_writing_interactive.py
    
    Or for help:
    python grade_ace_writing_interactive.py --help

AUTHOR: Educational Automation Library
VERSION: 2.0 (Interactive)
"""

import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from pathlib import Path

# ============================================================================
# HELP DOCUMENTATION
# ============================================================================

HELP_TEXT = """
╔══════════════════════════════════════════════════════════════════════════╗
║           ACE RUBRIC WRITING GRADER - INTERACTIVE VERSION                ║
╚══════════════════════════════════════════════════════════════════════════╝

PURPOSE:
    Automatically grades short answer writing responses using the ACE 
    (Answer, Cite, Explain) rubric. Provides scores for each component
    and includes students who didn't submit.

BEFORE YOU START - YOU'LL NEED:
    
    1. Master Roster Excel File
       └─ Contains all students across all periods
       └─ Created using: roster_management/create_master_roster.py
       └─ Required columns: Student Name, Period
    
    2. Google Form Response CSV Files (one per period)
       └─ Export from Google Forms: Responses tab → Spreadsheet icon → 
          File → Download → CSV
       └─ Must contain columns: Timestamp, Username (email), Question Response
    
    3. Assignment Information
       └─ The text passage students read
       └─ The question they answered

FILE FORMAT EXAMPLES:
    
    Master Roster:
        Student Name          | Grade | Student ID | Period | Course
        ----------------------------------------------------------------
        Smith, John           | 10    | 12345      | 1      | APCSP
        Garcia, Maria         | 11    | 12346      | 1      | APCSP
    
    Google Form CSV:
        Timestamp              | Username              | Question Response
        --------------------------------------------------------------------
        1/27/2025 8:30:00 AM  | john.smith@school.edu | The answer is...
        1/27/2025 8:31:00 AM  | maria.garcia@school  | According to the text...

GRADING RUBRIC (0-6 points total):
    
    Answer (0-2):  Does the response answer all parts of the question?
    Cite (0-2):    Does it cite evidence from the text?
    Explain (0-2): Does it explain how evidence supports the answer?

OUTPUT:
    Excel file with:
    - Individual scores (Answer, Cite, Explain)
    - Total score (0-6)
    - Color-coded: Green (5-6), Yellow (3-4), Red (0-2)
    - Response preview for spot-checking
    - Statistics by period and overall

WORKFLOW:
    1. Export Google Form responses to CSV
    2. Run this script
    3. Follow the prompts to enter file paths
    4. Script will grade all submissions
    5. Review output Excel file
    6. Spot-check borderline scores
    7. Transfer to gradebook

TIPS:
    - Test with one period first
    - Use Tab key for autocomplete on file paths
    - Press Enter to accept default values shown in [brackets]
    - You can drag-and-drop files into terminal for instant path

═══════════════════════════════════════════════════════════════════════════

Press Enter to continue or Ctrl+C to exit...
"""

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def print_header():
    """Print a nice header for the script."""
    print("\n" + "="*78)
    print("  ACE RUBRIC WRITING GRADER - Interactive Mode")
    print("="*78 + "\n")

def print_section(title):
    """Print a section divider."""
    print(f"\n{'─'*78}")
    print(f"  {title}")
    print('─'*78)

def get_input(prompt, default=None, required=True):
    """
    Get user input with optional default value.
    
    Args:
        prompt: Question to ask user
        default: Default value if user presses Enter
        required: Whether input is required
    
    Returns:
        User's input or default value
    """
    if default:
        full_prompt = f"{prompt} [{default}]: "
    else:
        full_prompt = f"{prompt}: "
    
    while True:
        value = input(full_prompt).strip()
        
        # Use default if provided and user pressed Enter
        if not value and default:
            return default
        
        # If required and no value, ask again
        if required and not value:
            print("  ⚠ This field is required. Please enter a value.")
            continue
        
        return value

def get_file_path(prompt, default=None, must_exist=True):
    """
    Get a file path from user with validation.
    
    Args:
        prompt: Question to ask user
        default: Default path
        must_exist: Whether file must already exist
    
    Returns:
        Valid file path
    """
    while True:
        path = get_input(prompt, default)
        
        # Expand ~ and resolve path
        path = os.path.expanduser(path)
        path = os.path.abspath(path)
        
        # Remove quotes if user wrapped path in quotes
        path = path.strip('"').strip("'")
        
        if must_exist and not os.path.exists(path):
            print(f"  ⚠ File not found: {path}")
            print(f"  Please check the path and try again.")
            continue
        
        return path

def get_yes_no(prompt, default='y'):
    """Get yes/no response from user."""
    default_text = 'Y/n' if default.lower() == 'y' else 'y/N'
    response = input(f"{prompt} [{default_text}]: ").strip().lower()
    
    if not response:
        response = default
    
    return response in ['y', 'yes']

# ============================================================================
# GRADING FUNCTION
# ============================================================================

def grade_answer(response, period, keywords_answer, keywords_cite, keywords_explain):
    """
    Grade a student response based on ACE rubric.
    
    Args:
        response: Student's written response
        period: Class period number
        keywords_answer: Keywords that indicate a complete answer
        keywords_cite: Citation phrases to look for
        keywords_explain: Explanation/connection words to look for
    
    Returns:
        dict: Scores for answer, cite, explain, and total
    """
    response_lower = response.lower()
    
    # ANSWER SCORING (0-2)
    keyword_count = sum(1 for keyword in keywords_answer if keyword.lower() in response_lower)
    
    if keyword_count >= len(keywords_answer) * 0.8:  # 80% of keywords present
        answer_score = 2
    elif keyword_count >= len(keywords_answer) * 0.4:  # 40% of keywords present
        answer_score = 1
    else:
        answer_score = 0
    
    # CITE SCORING (0-2)
    has_quotes = '"' in response or '"' in response or '"' in response
    has_cite_language = any(phrase in response_lower for phrase in keywords_cite)
    
    if has_quotes and has_cite_language:
        cite_score = 2
    elif has_quotes or has_cite_language:
        cite_score = 2  # Either method shows citation
    elif len(response) > 100:
        cite_score = 1
    else:
        cite_score = 0
    
    # EXPLAIN SCORING (0-2)
    has_explanation = any(phrase in response_lower for phrase in keywords_explain)
    sentences = [s for s in response.split('.') if len(s.strip()) > 10]
    has_structure = len(sentences) >= 3
    
    if has_explanation and has_structure:
        explain_score = 2
    elif has_explanation or has_structure:
        explain_score = 1
    else:
        explain_score = 0
    
    # Quality check
    if len(response) < 50:
        if answer_score > 0:
            answer_score = max(0, answer_score - 1)
    
    total_score = answer_score + cite_score + explain_score
    
    return {
        'answer_score': answer_score,
        'cite_score': cite_score,
        'explain_score': explain_score,
        'total_score': total_score
    }

# ============================================================================
# MAIN INTERACTIVE SCRIPT
# ============================================================================

def main():
    """Main interactive grading workflow."""
    
    # Check for help flag
    if '--help' in sys.argv or '-h' in sys.argv:
        print(HELP_TEXT)
        input()  # Wait for user to read
        return
    
    print_header()
    
    print("Welcome! This script will guide you through grading ACE rubric writing.")
    print("You can press Ctrl+C at any time to cancel.\n")
    
    if get_yes_no("Would you like to see detailed help before starting?", default='n'):
        print(HELP_TEXT)
        input("Press Enter to continue...")
    
    # ========================================================================
    # STEP 1: Get Master Roster
    # ========================================================================
    
    print_section("Step 1: Master Roster")
    print("First, I need your master roster with all students.")
    print("This should be an Excel file created by create_master_roster.py\n")
    
    default_roster = "./Master_Class_Roster.xlsx"
    master_roster_path = get_file_path(
        "Enter path to master roster Excel file",
        default=default_roster if os.path.exists(default_roster) else None
    )
    
    # Load and validate master roster
    try:
        master_roster = pd.read_excel(master_roster_path)
        required_cols = ['Student Name', 'Period']
        missing_cols = [col for col in required_cols if col not in master_roster.columns]
        
        if missing_cols:
            print(f"\n  ⚠ ERROR: Master roster is missing required columns: {missing_cols}")
            print(f"  Please make sure your roster has: Student Name, Period")
            return
        
        periods_available = sorted(master_roster['Period'].unique())
        print(f"\n  ✓ Master roster loaded successfully!")
        print(f"  ✓ Found {len(master_roster)} students across periods: {periods_available}")
        
    except Exception as e:
        print(f"\n  ⚠ ERROR loading master roster: {e}")
        return
    
    # ========================================================================
    # STEP 2: Get Google Form Response Files
    # ========================================================================
    
    print_section("Step 2: Google Form Response Files")
    print("Now I need the CSV files with student responses from Google Forms.")
    print("You'll need one CSV file per class period.\n")
    
    form_files = []
    
    for period in periods_available:
        print(f"\nPeriod {period}:")
        has_file = get_yes_no(f"  Do you have responses for Period {period}?", default='y')
        
        if has_file:
            default_name = f"./Week_1_Period_{period}_Responses.csv"
            file_path = get_file_path(
                f"  Enter path to Period {period} responses CSV",
                default=default_name if os.path.exists(default_name) else None
            )
            form_files.append((file_path, period))
        else:
            print(f"  Skipping Period {period}")
    
    if not form_files:
        print("\n  ⚠ No response files provided. Nothing to grade!")
        return
    
    print(f"\n  ✓ Will process {len(form_files)} period(s)")
    
    # ========================================================================
    # STEP 3: Get Assignment Information
    # ========================================================================
    
    print_section("Step 3: Assignment Grading Keywords")
    print("To grade accurately, I need to know what keywords to look for.\n")
    
    print("For the ANSWER component, what key concepts should appear?")
    print("(Enter keywords separated by commas)")
    print("Example: magma, lava, pressure, gases\n")
    
    answer_keywords = input("  Answer keywords: ").strip()
    keywords_answer = [k.strip() for k in answer_keywords.split(',')]
    
    print("\n  ✓ Will check for answer keywords:", keywords_answer)
    
    # Use standard citation and explanation keywords
    keywords_cite = [
        'author states', 'article states', 'passage', 'according to',
        'text states', 'paragraph', 'excerpt', 'states that', 'mentions',
        'the text', 'the article', 'the passage'
    ]
    
    keywords_explain = [
        'this shows', 'this proves', 'this means', 'therefore',
        'because', 'which causes', 'as a result', 'this demonstrates',
        'this explains', 'validates', 'supports', 'connected', 'since',
        'thus', 'hence', 'consequently'
    ]
    
    print("  ✓ Using standard CITE keywords (quotes, 'according to', etc.)")
    print("  ✓ Using standard EXPLAIN keywords (because, this shows, etc.)")
    
    # ========================================================================
    # STEP 4: Output Location
    # ========================================================================
    
    print_section("Step 4: Output Location")
    print("Where should I save the graded results?\n")
    
    default_output = "./Write_it_Wednesday_Grades.xlsx"
    output_path = get_file_path(
        "Enter output file path",
        default=default_output,
        must_exist=False
    )
    
    # Ensure .xlsx extension
    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'
    
    print(f"\n  ✓ Will save grades to: {output_path}")
    
    # ========================================================================
    # STEP 5: Confirm and Process
    # ========================================================================
    
    print_section("Step 5: Ready to Grade")
    print("\nSummary:")
    print(f"  • Master roster: {len(master_roster)} students")
    print(f"  • Response files: {len(form_files)} periods")
    print(f"  • Output file: {output_path}")
    print()
    
    if not get_yes_no("Ready to start grading?", default='y'):
        print("\n  Cancelled. No files were modified.")
        return
    
    # ========================================================================
    # GRADING PROCESS
    # ========================================================================
    
    print_section("Grading in Progress")
    
    all_grades = []
    
    for file_path, period in form_files:
        print(f"\nProcessing Period {period}...")
        
        try:
            # Read form responses
            with open(file_path, 'r') as f:
                reader = csv.DictReader(f)
                responses = list(reader)
            
            # Create dictionary of submissions
            submissions = {}
            for response in responses:
                email = response['Username'].strip().lower()
                answer_text = response[list(response.keys())[2]]  # Third column
                submissions[email] = answer_text
            
            print(f"  Found {len(submissions)} submission(s)")
            
            # Get students for this period
            period_students = master_roster[master_roster['Period'] == period]
            
            # Grade each student
            for _, student in period_students.iterrows():
                student_name = student['Student Name']
                
                # Match email to student
                if ',' in student_name:
                    last_name, first_name = student_name.split(',', 1)
                    last_name = last_name.strip().lower().replace(' ', '').replace('-', '')
                    first_name = first_name.strip().lower().replace(' ', '').replace('-', '')
                else:
                    parts = student_name.split()
                    first_name = parts[0].lower() if parts else ""
                    last_name = ' '.join(parts[1:]).lower().replace(' ', '').replace('-', '') if len(parts) > 1 else ""
                
                # Find matching submission
                found_submission = None
                matched_email = None
                
                for email, answer_text in submissions.items():
                    # Remove hyphens from email for comparison to match cleaned names
                    email_clean = email.lower().replace('-', '')
                    if first_name in email_clean and last_name in email_clean:
                        found_submission = answer_text
                        matched_email = email
                        break
                
                if found_submission:
                    # Grade the response
                    grades = grade_answer(
                        found_submission,
                        period,
                        keywords_answer,
                        keywords_cite,
                        keywords_explain
                    )
                    
                    all_grades.append({
                        'Period': period,
                        'Student Name': student_name,
                        'Email': matched_email,
                        'Answer Score': grades['answer_score'],
                        'Cite Score': grades['cite_score'],
                        'Explain Score': grades['explain_score'],
                        'Total Score': grades['total_score'],
                        'Response': found_submission[:100] + '...' if len(found_submission) > 100 else found_submission
                    })
                else:
                    # No submission
                    all_grades.append({
                        'Period': period,
                        'Student Name': student_name,
                        'Email': 'Not submitted',
                        'Answer Score': 0,
                        'Cite Score': 0,
                        'Explain Score': 0,
                        'Total Score': 0,
                        'Response': 'NO SUBMISSION'
                    })
            
            print(f"  ✓ Graded {len(period_students)} student(s)")
            
        except Exception as e:
            print(f"  ⚠ ERROR processing Period {period}: {e}")
            continue
    
    # ========================================================================
    # CREATE EXCEL OUTPUT
    # ========================================================================
    
    print("\nCreating Excel grade sheet...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    
    # Headers
    headers = ['Period', 'Student Name', 'Email', 'Answer (0-2)', 'Cite (0-2)',
               'Explain (0-2)', 'Total Score (0-6)', 'Response Preview']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data with color coding
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for row_idx, grade in enumerate(all_grades, start=2):
        ws.cell(row=row_idx, column=1, value=grade['Period'])
        ws.cell(row=row_idx, column=2, value=grade['Student Name'])
        ws.cell(row=row_idx, column=3, value=grade['Email'])
        ws.cell(row=row_idx, column=4, value=grade['Answer Score'])
        ws.cell(row=row_idx, column=5, value=grade['Cite Score'])
        ws.cell(row=row_idx, column=6, value=grade['Explain Score'])
        
        total_cell = ws.cell(row=row_idx, column=7, value=grade['Total Score'])
        
        # Color code
        if grade['Total Score'] >= 5:
            total_cell.fill = green_fill
        elif grade['Total Score'] >= 3:
            total_cell.fill = yellow_fill
        else:
            total_cell.fill = red_fill
        
        ws.cell(row=row_idx, column=8, value=grade['Response'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50
    
    # Save
    wb.save(output_path)
    
    # ========================================================================
    # SUMMARY STATISTICS
    # ========================================================================
    
    print_section("Grading Complete!")
    
    total_students = len(all_grades)
    total_submitted = sum(1 for g in all_grades if g['Total Score'] > 0)
    total_avg = sum(g['Total Score'] for g in all_grades) / total_students if total_students > 0 else 0
    
    print(f"\n📊 Overall Results:")
    print(f"  • Total students: {total_students}")
    print(f"  • Submitted: {total_submitted} ({total_submitted/total_students*100:.1f}%)")
    print(f"  • Average score: {total_avg:.2f}/6")
    
    print(f"\n📊 By Period:")
    for period in sorted(set(g['Period'] for g in all_grades)):
        period_grades = [g for g in all_grades if g['Period'] == period]
        avg = sum(g['Total Score'] for g in period_grades) / len(period_grades)
        submitted = sum(1 for g in period_grades if g['Total Score'] > 0)
        print(f"  Period {period}: {submitted}/{len(period_grades)} submitted, avg {avg:.2f}/6")
    
    print(f"\n💾 Grades saved to:")
    print(f"  {output_path}")
    
    print("\n" + "="*78)
    print("Next steps:")
    print("  1. Open the Excel file to review scores")
    print("  2. Spot-check responses, especially borderline scores (3-4)")
    print("  3. Make any manual adjustments needed")
    print("  4. Transfer scores to your gradebook")
    print("="*78 + "\n")

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Cancelled by user. No files were modified.")
        sys.exit(0)
    except Exception as e:
        print(f"\n  ⚠ ERROR: {e}")
        print(f"\n  If you need help, run: python {sys.argv[0]} --help")
        sys.exit(1)
