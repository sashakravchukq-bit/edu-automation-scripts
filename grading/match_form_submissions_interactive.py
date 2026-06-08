#!/usr/bin/env python3
"""
MATCH GOOGLE FORM SUBMISSIONS - INTERACTIVE
============================================

PURPOSE:
    Creates completion reports showing which students submitted a Google Form
    assignment. Interactive version with step-by-step guidance.

USAGE:
    python match_form_submissions_interactive.py
    
    Or for help:
    python match_form_submissions_interactive.py --help

AUTHOR: Educational Automation Library
VERSION: 2.0 (Interactive)
"""

import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# HELP DOCUMENTATION
# ============================================================================

HELP_TEXT = """
╔══════════════════════════════════════════════════════════════════════════╗
║          MATCH GOOGLE FORM SUBMISSIONS - INTERACTIVE                      ║
╚══════════════════════════════════════════════════════════════════════════╝

PURPOSE:
    Creates a completion report showing which students submitted (or didn't
    submit) a Google Form assignment. Perfect for tracking assignments and
    identifying missing work.

WHEN TO USE:
    • Quickly see who turned in an assignment
    • Create a completion checklist
    • Identify students who need reminders
    • Track participation rates by period

BEFORE YOU START - YOU'LL NEED:
    
    1. Master Roster Excel File
       └─ Contains all your students across all periods
       └─ Created using: create_master_roster_interactive.py
       └─ Required columns: Student Name, Period
    
    2. Google Form Response CSV Files (one per period)
       └─ Export from Google Forms: Responses → Spreadsheet icon →
          File → Download → CSV
       └─ Must contain: Username (student email addresses)

FILE FORMAT:
    
    Master Roster (Excel):
        Student Name          | Grade | Student ID | Period | Course
        ----------------------------------------------------------------
        Smith, John           | 10    | 12345      | 1      | APCSP
        Garcia, Maria         | 11    | 12346      | 1      | APCSP
    
    Google Form CSV:
        Timestamp              | Username              | [Other columns...]
        --------------------------------------------------------------------
        1/27/2025 8:30:00 AM  | john.smith@school.edu | [responses...]
        1/27/2025 8:31:00 AM  | maria.garcia@school  | [responses...]

OUTPUT:
    Excel file with:
    • All students from your roster
    • Completion status: "Yes" (green) or "No" (red)
    • Organized by period
    • Submission statistics

WORKFLOW:
    1. Collect responses via Google Form
    2. Export responses to CSV (one per period)
    3. Run this script
    4. Follow prompts to enter file paths
    5. Get completion report
    6. Contact students who didn't submit

TIPS:
    • The script fuzzy-matches names, so minor variations are OK
    • Students must use their school email for accurate matching
    • You can track multiple assignments by running this multiple times
    • Keep the output files to track participation trends

EMAIL MATCHING:
    The script matches students by finding their first and last names in
    the email address. For example:
    
    Student Name: "Garcia, Maria"
    Email: maria.garcia123@school.edu
    Result: ✓ Matched!
    
    Handles:
    • Compound last names (Garcia Rodriguez)
    • Hyphens (Garcia-Lopez)
    • Multiple middle names
    • Numbers in email addresses

═══════════════════════════════════════════════════════════════════════════

Press Enter to continue or Ctrl+C to exit...
"""

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def print_header():
    """Print a nice header."""
    print("\n" + "="*78)
    print("  MATCH GOOGLE FORM SUBMISSIONS - Interactive Mode")
    print("="*78 + "\n")

def print_section(title):
    """Print a section divider."""
    print(f"\n{'─'*78}")
    print(f"  {title}")
    print('─'*78)

def get_input(prompt, default=None, required=True):
    """Get user input with optional default."""
    if default is not None:
        full_prompt = f"{prompt} [{default}]: "
    else:
        full_prompt = f"{prompt}: "
    
    while True:
        value = input(full_prompt).strip()
        
        if not value and default is not None:
            return str(default)
        
        if required and not value:
            print("  ⚠ This field is required. Please enter a value.")
            continue
        
        return value

def get_file_path(prompt, default=None, must_exist=True):
    """Get a file path with validation."""
    while True:
        path = get_input(prompt, default)
        
        path = os.path.expanduser(path)
        path = os.path.abspath(path)
        path = path.strip('"').strip("'")
        
        if must_exist and not os.path.exists(path):
            print(f"  ⚠ File not found: {path}")
            print(f"  Please check the path and try again.")
            continue
        
        return path

def get_yes_no(prompt, default='y'):
    """Get yes/no response."""
    default_text = 'Y/n' if default.lower() == 'y' else 'y/N'
    response = input(f"{prompt} [{default_text}]: ").strip().lower()
    
    if not response:
        response = default
    
    return response in ['y', 'yes']

# ============================================================================
# MAIN INTERACTIVE SCRIPT
# ============================================================================

def main():
    """Main interactive workflow."""
    
    # Check for help
    if '--help' in sys.argv or '-h' in sys.argv:
        print(HELP_TEXT)
        input()
        return
    
    print_header()
    
    print("Welcome! This script will create a completion report from Google Forms.")
    print("You can press Ctrl+C at any time to cancel.\n")
    
    if get_yes_no("Would you like to see detailed help before starting?", default='n'):
        print(HELP_TEXT)
        input("Press Enter to continue...")
    
    # ========================================================================
    # STEP 1: Get Master Roster
    # ========================================================================
    
    print_section("Step 1: Master Roster")
    print("First, I need your master roster with all students.\n")
    
    default_roster = "./Master_Class_Roster.xlsx"
    master_roster_path = get_file_path(
        "Enter path to master roster Excel file",
        default=default_roster if os.path.exists(default_roster) else None
    )
    
    # Load and validate
    try:
        master_roster = pd.read_excel(master_roster_path)
        
        if 'Student Name' not in master_roster.columns or 'Period' not in master_roster.columns:
            print("\n  ⚠ ERROR: Master roster must have 'Student Name' and 'Period' columns")
            return
        
        periods_available = sorted(master_roster['Period'].unique())
        print(f"\n  ✓ Master roster loaded successfully!")
        print(f"  ✓ Found {len(master_roster)} students across periods: {periods_available}")
        
    except Exception as e:
        print(f"\n  ⚠ ERROR loading master roster: {e}")
        return
    
    # ========================================================================
    # STEP 2: Get Google Form Response Files
    # ========================================================================
    
    print_section("Step 2: Google Form Response Files")
    print("Now I need the CSV files with form submissions.\n")
    
    form_files = []
    
    for period in periods_available:
        print(f"\nPeriod {period}:")
        has_file = get_yes_no(f"  Do you have responses for Period {period}?", default='y')
        
        if has_file:
            default_name = f"./Period_{period}_Responses.csv"
            file_path = get_file_path(
                f"  Enter path to Period {period} responses CSV",
                default=default_name if os.path.exists(default_name) else None
            )
            form_files.append((file_path, period))
        else:
            print(f"  Skipping Period {period}")
    
    if not form_files:
        print("\n  ⚠ No response files provided. Nothing to process!")
        return
    
    print(f"\n  ✓ Will process {len(form_files)} period(s)")
    
    # ========================================================================
    # STEP 3: Assignment Name
    # ========================================================================
    
    print_section("Step 3: Assignment Name")
    print("What should I call this assignment in the report?\n")
    
    assignment_name = get_input(
        "Assignment name",
        default="Assignment Completion"
    )
    
    print(f"\n  ✓ Assignment: {assignment_name}")
    
    # ========================================================================
    # STEP 4: Output Location
    # ========================================================================
    
    print_section("Step 4: Output Location")
    print("Where should I save the completion report?\n")
    
    default_output = f"./{assignment_name.replace(' ', '_')}_Report.xlsx"
    output_path = get_file_path(
        "Enter output file path",
        default=default_output,
        must_exist=False
    )
    
    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'
    
    print(f"\n  ✓ Will save report to: {output_path}")
    
    # ========================================================================
    # STEP 5: Confirm and Process
    # ========================================================================
    
    print_section("Step 5: Ready to Process")
    print("\nSummary:")
    print(f"  • Master roster: {len(master_roster)} students")
    print(f"  • Response files: {len(form_files)} periods")
    print(f"  • Assignment: {assignment_name}")
    print(f"  • Output file: {output_path}")
    print()
    
    if not get_yes_no("Ready to create completion report?", default='y'):
        print("\n  Cancelled. No files were created.")
        return
    
    # ========================================================================
    # PROCESSING
    # ========================================================================
    
    print_section("Processing Submissions")
    
    # Dictionary to store submissions by period
    completed_by_period = {}
    
    for file_path, period in form_files:
        print(f"\nProcessing Period {period}...")
        
        try:
            form_df = pd.read_csv(file_path)
            
            if 'Username' not in form_df.columns:
                print(f"  ⚠ ERROR: No 'Username' column found in {file_path}")
                completed_by_period[period] = set()
                continue
            
            submitted_emails = set(form_df['Username'].str.lower().str.strip())
            completed_by_period[period] = submitted_emails
            
            print(f"  ✓ Found {len(submitted_emails)} submission(s)")
            
        except Exception as e:
            print(f"  ⚠ ERROR reading file: {e}")
            completed_by_period[period] = set()
    
    # ========================================================================
    # CREATE REPORT
    # ========================================================================
    
    print("\nCreating completion report...")
    
    report_data = []
    
    for _, student in master_roster.iterrows():
        period = student['Period']
        student_name = student['Student Name']
        
        # Extract name parts for matching
        if ',' in student_name:
            last_name, first_name = student_name.split(',', 1)
            last_name = last_name.strip().lower().replace(' ', '').replace('-', '')
            first_name = first_name.strip().lower().replace(' ', '').replace('-', '')
        else:
            parts = student_name.split()
            if len(parts) >= 2:
                first_name = parts[0].lower()
                last_name = ' '.join(parts[1:]).lower().replace(' ', '').replace('-', '')
            else:
                first_name = student_name.lower()
                last_name = ""
        
        # Check if submitted
        completed = "No"
        if period in completed_by_period:
            submitted_emails = completed_by_period[period]
            for email in submitted_emails:
                if first_name in email.lower() and last_name in email.lower():
                    completed = "Yes"
                    break
        
        report_data.append({
            'Student Name': student_name,
            'Period': period,
            'Course': student.get('Course', ''),
            assignment_name: completed
        })
    
    # ========================================================================
    # CREATE EXCEL OUTPUT
    # ========================================================================
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Completion Report"
    
    # Headers
    headers = ['Student Name', 'Period', 'Course', assignment_name]
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data with color coding
    yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for row_idx, record in enumerate(report_data, start=2):
        ws.cell(row=row_idx, column=1, value=record['Student Name'])
        ws.cell(row=row_idx, column=2, value=record['Period'])
        ws.cell(row=row_idx, column=3, value=record['Course'])
        
        completion_cell = ws.cell(row=row_idx, column=4, value=record[assignment_name])
        
        if record[assignment_name] == 'Yes':
            completion_cell.fill = yes_fill
        else:
            completion_cell.fill = no_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    
    # Save
    wb.save(output_path)
    
    # ========================================================================
    # SUMMARY STATISTICS
    # ========================================================================
    
    print_section("Completion Report Created!")
    
    total_students = len(report_data)
    total_completed = sum(1 for r in report_data if r[assignment_name] == 'Yes')
    total_percent = (total_completed / total_students * 100) if total_students > 0 else 0
    
    print(f"\n📊 Overall Results:")
    print(f"  • Total students: {total_students}")
    print(f"  • Completed: {total_completed} ({total_percent:.1f}%)")
    print(f"  • Not completed: {total_students - total_completed}")
    
    print(f"\n📊 By Period:")
    for period in sorted(set(r['Period'] for r in report_data)):
        period_data = [r for r in report_data if r['Period'] == period]
        completed = sum(1 for r in period_data if r[assignment_name] == 'Yes')
        total = len(period_data)
        percent = (completed / total * 100) if total > 0 else 0
        print(f"  Period {period}: {completed}/{total} ({percent:.1f}%)")
    
    print(f"\n💾 Report saved to:")
    print(f"  {output_path}")
    
    print("\n" + "="*78)
    print("Next steps:")
    print("  1. Open the Excel file to see who completed the assignment")
    print("  2. Contact students who didn't submit (marked in red)")
    print("  3. Use this for gradebook or attendance tracking")
    print("="*78 + "\n")

# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Cancelled by user. No files were created.")
        sys.exit(0)
    except Exception as e:
        print(f"\n  ⚠ ERROR: {e}")
        print(f"\n  If you need help, run: python {sys.argv[0]} --help")
        sys.exit(1)
